#!/usr/bin/env python
# -*- coding: utf-8 -*-
USAGE = """convert_to_csv.py

Converts the Military Surplus XLSX to an aggregate CSV file as well as a
directory of CSV files by state.

Usage:
    convert_to_csv.py to_file <source> <destination>

"""
import sys
import csv

from docopt import docopt
from openpyxl import load_workbook


def _to_file(wb, writer):
    for sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(name=sheet_name)
        sheet_i = sheet.iter_rows()
        # Skip over the header row.
        next(sheet_i)

        for row in sheet_i:
            writer.writerow([cell.value for cell in row])


def main():
    args = docopt(USAGE)
    wb = load_workbook(filename=args['<source>'], use_iterators=True)

    if args['to_file']:
        with open(args['<destination>'], 'wb') as fout:
            writer = csv.writer(fout)
            writer.writerow([
                cell.value
                for cell in next(wb.active.iter_rows())
            ])
            return _to_file(wb, writer)


if __name__ == '__main__':
    sys.exit(main())
